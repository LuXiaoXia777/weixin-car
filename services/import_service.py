"""解析微信后台导出的 CSV/XLSX，并转换成统一导入模型。"""

from __future__ import annotations

import csv
import hashlib
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from database.models import ArticleImportRow, ImportBatch, UserImportRow


ALIASES = {
    "title": ("标题", "文章标题", "图文标题", "title"),
    "url": ("链接", "文章链接", "图文链接", "url"),
    "digest": ("摘要", "文章摘要", "digest"),
    "category": ("分类", "内容类型", "category"),
    "external_id": ("文章id", "图文id", "article_id", "external_id"),
    "publish_time": ("发布时间", "发表时间", "publish_time"),
    "stat_date": ("日期", "数据日期", "统计日期", "date", "stat_date", "ref_date"),
    "views": ("阅读", "阅读量", "阅读次数", "图文阅读次数", "views", "read_count"),
    "read_users": ("阅读人数", "read_users", "read_user"),
    "likes": ("点赞", "点赞数", "点赞人数", "likes", "like_user"),
    "recommendations": ("在看", "在看数", "爱心赞", "recommendations", "zaikan_user"),
    "shares": ("分享", "分享数", "分享次数", "转发", "转发次数", "shares", "share_user"),
    "comments": ("评论", "评论数", "留言", "留言数", "comments", "comment_count"),
    "new_followers": ("新增粉丝", "新增关注", "新增用户", "new_followers", "new_user"),
    "cancel_followers": ("取消粉丝", "取消关注", "取消用户", "cancel_followers", "cancel_user"),
    "net_followers": ("净增粉丝", "净增关注", "净增用户", "net_followers"),
}


def _normalize_header(value: object) -> str:
    return re.sub(r"[\s_\-（）()]+", "", str(value or "").strip().lower())


HEADER_LOOKUP = {
    _normalize_header(alias): canonical
    for canonical, aliases in ALIASES.items()
    for alias in aliases
}


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _integer(value: Any, field: str) -> int | None:
    if value is None or str(value).strip() in ("", "-", "--", "—"):
        return None
    cleaned = str(value).replace(",", "").strip()
    try:
        number = float(cleaned)
    except ValueError as exc:
        raise ValueError(f"字段 {field} 不是有效整数：{value}") from exc
    if not number.is_integer():
        raise ValueError(f"字段 {field} 必须是整数：{value}")
    return int(number)


def _datetime(value: Any, field: str) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"字段 {field} 不是支持的日期格式：{value}")


def _source_key(title: str, publish_time: datetime, url: str | None) -> str:
    if url:
        return f"url:{url.strip()}"
    digest = hashlib.sha256(f"{title.strip()}|{publish_time.isoformat()}".encode()).hexdigest()
    return f"title_time:{digest}"


def _read_csv(path: Path) -> list[tuple[str, list[list[Any]]]]:
    content = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            content = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if content is None:
        raise ValueError("CSV 编码无法识别，请保存为 UTF-8 或 GB18030")
    return [(path.stem, [list(row) for row in csv.reader(content.splitlines())])]


def _read_xlsx(path: Path) -> list[tuple[str, list[list[Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if any(any(value not in (None, "") for value in row) for row in rows):
                tables.append((sheet.title, rows))
    finally:
        workbook.close()
    return tables


def _header_row(rows: list[list[Any]]) -> tuple[int, dict[int, str]]:
    for index, row in enumerate(rows[:20]):
        mapping = {}
        for column, value in enumerate(row):
            canonical = HEADER_LOOKUP.get(_normalize_header(value))
            if canonical and canonical not in mapping.values():
                mapping[column] = canonical
        has_identity = "title" in mapping.values() or any(
            name in mapping.values() for name in ("new_followers", "cancel_followers", "net_followers")
        )
        if len(mapping) >= 2 and has_identity:
            return index, mapping
    raise ValueError("前20行中未找到可识别的表头")


def _records(rows: list[list[Any]]) -> Iterable[dict[str, Any]]:
    header_index, mapping = _header_row(rows)
    for row in rows[header_index + 1 :]:
        record = {name: row[column] if column < len(row) else None for column, name in mapping.items()}
        if any(value not in (None, "") for value in record.values()):
            yield record


def _article_row(record: dict[str, Any]) -> ArticleImportRow:
    title = _text(record.get("title"))
    if not title:
        raise ValueError("文章行缺少标题")
    publish_raw = record.get("publish_time") or record.get("stat_date")
    publish_time = _datetime(publish_raw, "发布时间")
    stat_raw = record.get("stat_date") or publish_time
    stat_date = _datetime(stat_raw, "统计日期").date()
    url = _text(record.get("url"))
    return ArticleImportRow(
        source_key=_source_key(title, publish_time, url),
        title=title,
        publish_time=publish_time,
        stat_date=stat_date,
        url=url,
        digest=_text(record.get("digest")),
        category=_text(record.get("category")),
        external_id=_text(record.get("external_id")),
        views=_integer(record.get("views"), "阅读量"),
        read_users=_integer(record.get("read_users"), "阅读人数"),
        likes=_integer(record.get("likes"), "点赞"),
        recommendations=_integer(record.get("recommendations"), "在看"),
        shares=_integer(record.get("shares"), "分享"),
        comments=_integer(record.get("comments"), "评论"),
        new_followers=_integer(record.get("new_followers"), "文章涨粉"),
    )


def _user_row(record: dict[str, Any]) -> UserImportRow:
    stat_date = _datetime(record.get("stat_date"), "统计日期").date()
    new = _integer(record.get("new_followers"), "新增关注")
    cancel = _integer(record.get("cancel_followers"), "取消关注")
    net = _integer(record.get("net_followers"), "净增关注")
    if net is None and new is not None and cancel is not None:
        net = new - cancel
    return UserImportRow(stat_date=stat_date, new_followers=new, cancel_followers=cancel, net_followers=net)


def parse_import_file(path: Path) -> ImportBatch:
    path = path.expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"找不到导入文件：{path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        tables = _read_csv(path)
    elif suffix == ".xlsx":
        tables = _read_xlsx(path)
    else:
        raise ValueError("仅支持 .csv 和 .xlsx 文件")

    articles: list[ArticleImportRow] = []
    user_stats: list[UserImportRow] = []
    errors = []
    for sheet_name, rows in tables:
        try:
            records = list(_records(rows))
        except ValueError as exc:
            errors.append(f"{sheet_name}: {exc}")
            continue
        for row_number, record in enumerate(records, start=2):
            try:
                if _text(record.get("title")):
                    articles.append(_article_row(record))
                elif any(record.get(key) not in (None, "") for key in ("new_followers", "cancel_followers", "net_followers")):
                    user_stats.append(_user_row(record))
            except ValueError as exc:
                errors.append(f"{sheet_name} 第{row_number}行: {exc}")

    if errors:
        raise ValueError("导入文件校验失败：\n" + "\n".join(errors[:20]))
    if not articles and not user_stats:
        raise ValueError("导入文件中没有可识别的文章或用户数据")
    return ImportBatch(
        source_file=path,
        file_hash=hashlib.sha256(path.read_bytes()).hexdigest(),
        articles=articles,
        user_stats=user_stats,
    )
